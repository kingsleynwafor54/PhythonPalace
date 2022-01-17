from openpyxl import Workbook, load_workbook
print("hello world")
wb=load_workbook("grades/test_csv_DB.xlsx")
ws=wb.active
print(ws['A1'].value)